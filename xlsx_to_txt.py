import openpyxl
from difflib import SequenceMatcher
import sys
import os
import subprocess

def main():
    script_dir = os.path.dirname(__file__) #<-- absolute dir the script is in
    # sheet_rel_path = "sheet.xlsx"
    sheet_rel_path = "Book4.xlsx"
    sheet_abs_file_path = os.path.join(script_dir, sheet_rel_path)
    output_rel_path = "interim/sheet_contents.txt"
    output_abs_file_path = os.path.join(script_dir, output_rel_path)

    try:
        wrkbk = openpyxl.load_workbook(sheet_abs_file_path)
    except OSError:
       print("Could not open/read file: ", sheet_rel_path)
       sys.exit()
    
    output = open(output_abs_file_path, "w+")
    sh = wrkbk.active

    # INIT LISTS
    # list_count = int(sys.argv[1])
    list_count = 2 # is this always true?
    lists = []
    for i in range(list_count):
        lists.append([])
    # INIT LISTS

    # PARSE XLSX
    list_i = 0      # list_i keeps track of which list we're appending, necessary to handle different file formats
    for i in range(1, sh.max_column+1):                # range starts from 1 because excel sheets start from 1
        if sh.cell(row=1, column=i).value is None and sh.cell(row=2, column=i).value is None:   # if 1st and 2nd elements of the i-th column are empty, assume the whole column is empty
            continue
        for j in range(1, sh.max_row+1):
            cell_obj = sh.cell(row=j, column=i)
            cell = str(cell_obj.value)
            if cell == "Grand Total" or "list" in cell or "List" in cell or "LIST" in cell: # so we don't include list headers
                    continue
            elif cell != "None":            # "None" is the string value xlsx gives to empty cells
                lists[list_i].append(cell)
        list_i += 1
    # PARSE XLSX

    for l in lists:
        for j in l:
            output.write(j)
            output.write("\n")
        output.write("\n")
    output.close()

    subprocess.run(["bin/sort"])        # run compiled C++ code

main()